from openpyxl import Workbook, load_workbook
import datetime
import re

file_path = 'C:/Job/일일업무_내용_정리_20190510.xlsx'
wb = load_workbook(file_path)
print(wb.sheetnames)
#wb.save("C:/Job/sample.xlsx")
#wb.close()
#ws = wb.active
ws = wb['2분기']
print(ws.cell(row = 2, column = 1).value)
row_cnt = ws.max_row
col_cnt = ws.max_column
print('rows: %d , cols: %d' % (row_cnt, col_cnt))
#ws1 = wb.copy_worksheet(ws) # worksheet copy


# 줄 사이에 빈 줄 제거
pattern = re.compile(r'\n+')
for y in range(1, row_cnt):
    row_idx = y+1
    org_value = ws.cell(row = row_idx, column = 2).value
    if org_value:
        match = re.sub(pattern, '\n', org_value) # 빈 줄 제거
        # 원본Cell 옆으로 수정본Cell 추가
        #ws.cell(row=row_idx, column=3).value = match
        # 원본Cell에 바로 수정 반영
        ws.cell(row=row_idx, column=2).value = match
    else:
        #print('cell[{0}:{1}] : '.format(row_idx, 2, org_value))
        continue

    #print(match)

save_file_path = "C:/Job/일일업무_내용_정리_{:%Y%m%d%H%M%S}.xlsx".format(datetime.datetime.now())  # '_20190503135732'
wb.save(save_file_path)
wb.close()