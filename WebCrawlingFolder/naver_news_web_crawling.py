# -*- coding: utf-8 -*-
import requests
from bs4 import BeautifulSoup
import re
import time
from openpyxl import Workbook
#from openpyxl.styles import Font, Fill
import datetime

"""
(venv_name)command>pip installer pyinstaller
# exe file make
(venv_name)command> python -O -m PyInstaller --onefile naver_news_web_crawling.py
"""

class xitem:
    def __init__(self, title, link, source, term):
        self.title = title
        self.link = link
        self.source = source
        self.term = term

    def __str__(self):
        return 'title:' + self.title + ', link:' + self.link + ', source:' + self.source + ', term:' + str(self.term)


def get_news_items(url):
    xitem_list = []
    res = requests.get(url)
    obj = BeautifulSoup(res.text, 'html.parser')
    lst = obj.find('ul', {'class': 'type01'})
    lst_a = lst.find_all('a', {'class': '_sp_each_title'})  # lst.findAll('a', attrs='_sp_each_title')

    for x in lst_a:
        #    print(x)
        title = x['title']
        link = x['href']
        p1 = x.find_parent().find_parent()
        p2 = p1.find('dd', {'class': 'txt_inline'})
        source = p2.find('span', {'class': '_sp_each_source'}).get_text()
        term = p2.find(string=re.compile(r'([0-9]+(분 전|시간 전|일 전)) | (\d{4}.\d{2}.\d{2})'))

        xitem_list.append(xitem(title, link, source, term))

    return xitem_list


def adjust_column_width_auto(ws):
    """ 각 열들의 너비 자동 조정 """
    column_widths = []
    for row in ws.iter_rows():
        for i, cell in enumerate(row):
            try:
                column_widths[i] = max(column_widths[i], len(cell.value))
            except IndexError:
                column_widths.append(len(cell.value))
    # print(column_widths)
    for i, column_width in enumerate(column_widths):
        column_name = chr(65 + i)
        # print(column_name)
        ws.column_dimensions[column_name].width = column_width


def excel_file_save(file_name, data_list):
    wb = Workbook()
    sheet1 = wb.active
    # sheet1.title = 'Sheet1'
    min_row = 1
    max_row = len(data_list) + 1
    # max_col = 4
    for row_idx in range(min_row, max_row):
        data = data_list[row_idx - 1]  # 0부터 시작
        # print(data)

        # col_1_length if col_1_length > len(data.title) else len(data.title)

        sheet1.cell(column=1, row=row_idx, value="{0}".format(data.title))
        sheet1.cell(column=1, row=row_idx).hyperlink = data.link
        sheet1.cell(column=1, row=row_idx).style = 'Hyperlink'
        sheet1.cell(column=2, row=row_idx, value="{0}".format(data.source))
        sheet1.cell(column=3, row=row_idx, value="{0}".format(data.term))

        # column width change
    adjust_column_width_auto(sheet1)

    wb.save(file_name)


def input_data():
    # cur_dir = os.getcwd()
    # print('current directory : ' + cur_dir)
    print('program exit : \'x\' press key or \'Enter\' key')
    while True:
        input_data = input('input data : ')
        if input_data == 'x' or input_data == '':
            break
        else:
            if 0 < len(input_data):
                return input_data
    return ''


def main():
    search_text = input_data()
    if len(search_text) <= 0:
        print('program exit.')
        return

    # url_format = 'https://search.naver.com/search.naver?&where=news&query={0}&sm=tab_pge&sort=1&photo=0&field=0&reporter_article=&pd=0&ds=&de=&docid=&nso=so:dd,p:all,a:all&mynews=0&start={1}&refresh_start=1'
    #url_format = 'https://search.naver.com/search.naver?where=news&query={0}&sm=tab_srt&sort=1&photo=0&field=0&reporter_article=&pd=0&ds=&de=&docid=&nso=so%3Add%2Cp%3Aall%2Ca%3Aall&mynews=0&refresh_start={1}&related=0'
    url_format = 'https://search.naver.com/search.naver?query={0}&where=news&ie=utf8&sm=tab_pge&sort=1&start={1}'
    xitem_list = []

    print('검색을 시작합니다.')
    for i in range(0, 10, 1):  # 1.11.21.31.41.51.61.71.81.91
        print(str(i + 1), 'Page ', end=" ")
        # print('index : ' + str(i))
        url = url_format.format(search_text, i * 10 + 1)

        xitem_list += get_news_items(url)

        time.sleep(1)

    print("***" + str(len(xitem_list)) + ' 건이 검색되었습니다.')
    for x in xitem_list:
        print(x)

    # print(xitem_list[0])
    save_file_name = "save_sample_{:%Y%m%d%H%M%S}.xlsx".format(datetime.datetime.now())
    excel_file_save(save_file_name, xitem_list)

    print('Good By!')
    return

if __name__ == '__main__':
    main()